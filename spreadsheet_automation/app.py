import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#load workbook from directory
def excel_processor(filename):
    wb = xl.load_workbook(filename)

    # Grab excel sheet to work with.
    # In order to know the available sheet names,
    # use the wb.sheetnames to view sheet arrays
    sheet = wb['Sheet1']
    sheet['D1'] = "Corrected Price"

    # Access sheet cell
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price


    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.x_axis.title = "New Price"
    chart.y_axis.title = "Y-Axis"
    chart.title = "New corrected price chart"
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')

    wb.save(filename)


excel_processor('transactions.xlsx')